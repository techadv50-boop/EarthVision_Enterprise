"""Per-site and master report generation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from webcrawler.crawler.site_crawler import SiteResult

MASTER_HEADERS = [
    "Website",
    "Total Pages",
    "PDFs",
    "Word Files",
    "Excel Files",
    "PowerPoint Files",
    "Images",
    "Emails",
    "Phone Numbers",
    "Start Time",
    "Finish Time",
    "Processing Time",
    "Status",
]


def write_summary(result: SiteResult) -> Path:
    assert result.site_dir is not None
    path = result.site_dir / "Reports" / "Summary.txt"
    path.parent.mkdir(parents=True, exist_ok=True)
    content = f"""Website URL: {result.website}
Total Pages Crawled: {result.pages_crawled}
Total Documents Downloaded / Scanned: {result.documents_downloaded}
PDFs: {result.pdfs}
Word Files: {result.word_files}
Excel Files: {result.excel_files}
PowerPoint Files: {result.powerpoint_files}
Images: {result.images}
Emails Extracted: {result.emails}
Phone Numbers Extracted: {result.phones}
Start Time: {result.start_time}
End Time: {result.end_time}
Total Processing Time: {result.processing_time}
Status: {result.status}
Notes: emails.txt and phone_numbers.txt are written in the site folder.
"""
    if result.error:
        content += f"Error: {result.error}\n"
    path.write_text(content, encoding="utf-8")
    return path


def append_master_report(output_root: Path | str, result: SiteResult) -> Path:
    root = Path(output_root)
    root.mkdir(parents=True, exist_ok=True)
    path = root / "Master_Report.xlsx"
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Report"
        ws.append(MASTER_HEADERS)

    ws.append(
        [
            result.website,
            result.pages_crawled,
            result.pdfs,
            result.word_files,
            result.excel_files,
            result.powerpoint_files,
            result.images,
            result.emails,
            result.phones,
            result.start_time,
            result.end_time,
            result.processing_time,
            result.status,
        ]
    )
    wb.save(path)
    return path
