"""Extract article metadata from DOIs via landing-page scrape + Crossref fallback."""

from __future__ import annotations

import asyncio
import re
from html import unescape
from typing import Iterable

import httpx
from bs4 import BeautifulSoup

from .models import ArticleMeta, Author, FileLink
from .redif import month_name, normalize_doi_url

USER_AGENT = "DOI-REDIF-Converter/1.0 (mailto:redif-converter@local; academic metadata extraction)"
MONTH_MAP = {
    "january": "1",
    "february": "2",
    "march": "3",
    "april": "4",
    "may": "5",
    "june": "6",
    "july": "7",
    "august": "8",
    "september": "9",
    "october": "10",
    "november": "11",
    "december": "12",
}


def clean_doi(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    value = value.strip("<>[]()'\"")
    value = re.sub(r"^https?://(dx\.)?doi\.org/", "", value, flags=re.I)
    value = value.split("?")[0].split("#")[0].strip().rstrip("/")
    # Strip trailing punctuation commonly copied from citations
    value = value.rstrip(".,;:)")
    # Keep only plausible DOI shapes
    if not re.match(r"^10\.\d{4,9}/\S+$", value):
        return ""
    return value


def clean_url(value: str) -> str:
    value = (value or "").strip().strip("<>[]()'\"")
    if not value:
        return ""
    value = value.split()[0].strip().rstrip(".,);]")
    if not re.match(r"^https?://", value, flags=re.I):
        return ""
    return value


def is_doi_org_url(value: str) -> bool:
    return bool(re.search(r"https?://(dx\.)?doi\.org/", value or "", flags=re.I))


def parse_doi_list(text: str) -> list[str]:
    """Parse DOIs from pasted text (one per line, comma, or whitespace separated)."""
    parts = re.split(r"[\s,;]+", text or "")
    seen: set[str] = set()
    out: list[str] = []
    for part in parts:
        doi = clean_doi(part)
        if not doi or "10." not in doi:
            continue
        # keep first occurrence, case-preserving
        key = doi.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(doi)
    return out


def parse_input_list(text: str) -> list[str]:
    """Parse DOIs and/or article URLs from pasted text.

    Prefer one item per line. Article URLs (http/https) are kept as-is.
    DOI links and bare DOIs are normalized to DOI strings.
    """
    seen: set[str] = set()
    out: list[str] = []

    def add(item: str) -> None:
        key = item.lower()
        if key in seen:
            return
        seen.add(key)
        out.append(item)

    raw = text or ""
    lines = raw.splitlines()
    if len(lines) <= 1 and re.search(r"https?://", raw, flags=re.I):
        # Split URL-heavy blobs carefully
        chunks = re.split(r"[\n\r]+|,\s*(?=https?://)|;\s*(?=https?://)", raw)
    else:
        chunks = lines if lines else [raw]

    for chunk in chunks:
        chunk = (chunk or "").strip().strip(",;")
        if not chunk:
            continue
        url = clean_url(chunk)
        if url:
            if is_doi_org_url(url):
                doi = clean_doi(url)
                if doi:
                    add(doi)
            else:
                add(url)
            continue
        for doi in parse_doi_list(chunk):
            add(doi)
    return out


def inputs_from_xlsx_bytes(data: bytes) -> list[str]:
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    seen: set[str] = set()
    out: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                for item in parse_input_list(str(cell)):
                    key = item.lower()
                    if key in seen:
                        continue
                    seen.add(key)
                    out.append(item)
    wb.close()
    return out


def dois_from_xlsx_bytes(data: bytes) -> list[str]:
    """Backward-compatible alias: returns DOIs and article URLs."""
    return inputs_from_xlsx_bytes(data)


def _strip_jats(text: str) -> str:
    if not text:
        return ""
    text = unescape(text)
    text = re.sub(r"</?jats:[^>]+>", "", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _meta_values(soup: BeautifulSoup, *names: str) -> list[str]:
    values: list[str] = []
    for name in names:
        for tag in soup.find_all("meta", attrs={"name": name}):
            content = (tag.get("content") or "").strip()
            if content:
                values.append(content)
        for tag in soup.find_all("meta", attrs={"property": name}):
            content = (tag.get("content") or "").strip()
            if content:
                values.append(content)
    return values


def _first(*values: Iterable[str] | str | None) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return value.strip()
            continue
        for item in value:
            if item and str(item).strip():
                return str(item).strip()
    return ""


def _pdf_view_url(pdf_url: str) -> str:
    """Prefer OJS /article/view/.../galley style used in sample REDIF."""
    if "/article/download/" in pdf_url:
        return pdf_url.replace("/article/download/", "/article/view/")
    return pdf_url


def _parse_authors_from_citation(soup: BeautifulSoup) -> list[Author]:
    authors: list[Author] = []
    current: Author | None = None
    for tag in soup.find_all("meta"):
        name = (tag.get("name") or "").lower()
        content = (tag.get("content") or "").strip()
        if not content:
            continue
        if name == "citation_author":
            if current:
                authors.append(current)
            current = Author(name=re.sub(r"\s+", " ", content))
        elif name == "citation_author_institution" and current is not None:
            if current.workplace:
                current.workplace = f"{current.workplace}; {content}"
            else:
                current.workplace = content
        elif name == "citation_author_email" and current is not None:
            current.email = content
    if current:
        authors.append(current)

    # Fallback: DC.Creator.PersonalName
    if not authors:
        for name in _meta_values(soup, "DC.Creator.PersonalName", "dc.creator"):
            authors.append(Author(name=re.sub(r"\s+", " ", name)))
    return authors


def _extract_from_html(html: str, landing_url: str, doi: str) -> ArticleMeta:
    soup = BeautifulSoup(html, "lxml")
    authors = _parse_authors_from_citation(soup)

    title = _first(
        _meta_values(soup, "citation_title", "DC.Title", "dc.title", "og:title"),
    )
    abstract = _first(
        _meta_values(soup, "citation_abstract", "DC.Description", "dc.description", "og:description"),
    )
    keywords = _meta_values(soup, "citation_keywords", "DC.Subject", "dc.subject")
    # dedupe keywords preserving order
    seen_kw: set[str] = set()
    clean_keywords: list[str] = []
    for kw in keywords:
        for part in re.split(r"[;,]", kw):
            part = part.strip()
            key = part.lower()
            if part and key not in seen_kw:
                seen_kw.add(key)
                clean_keywords.append(part)

    journal = _first(
        _meta_values(soup, "citation_journal_title", "DC.Source", "dc.source"),
    )
    # Normalize ampersand style to match sample ("and")
    journal = journal.replace(" & ", " and ")

    volume = _first(_meta_values(soup, "citation_volume", "DC.Source.Volume"))
    issue = _first(_meta_values(soup, "citation_issue", "DC.Source.Issue"))
    first = _first(_meta_values(soup, "citation_firstpage"))
    last = _first(_meta_values(soup, "citation_lastpage"))
    pages = _first(_meta_values(soup, "DC.Identifier.pageNumber", "citation_pages"))
    if not pages and first:
        pages = f"{first}-{last}" if last and last != first else first

    date = _first(
        _meta_values(
            soup,
            "citation_publication_date",
            "citation_date",
            "citation_online_date",
            "DC.Date.issued",
            "dc.date",
        )
    )
    year = ""
    month = ""
    if date:
        # YYYY/MM/DD or YYYY-MM-DD
        m = re.match(r"(\d{4})(?:[/-](\d{1,2}))?", date)
        if m:
            year = m.group(1)
            if m.group(2):
                month = str(int(m.group(2)))
        else:
            for name, num in MONTH_MAP.items():
                if name in date.lower():
                    month = num
                    break
            y = re.search(r"(19|20)\d{2}", date)
            if y:
                year = y.group(0)

    pdf_urls = _meta_values(soup, "citation_pdf_url")
    html_url = _first(
        _meta_values(soup, "citation_abstract_html_url", "DC.Identifier.URI"),
        landing_url,
    )

    file_links: list[FileLink] = []
    if pdf_urls:
        file_links.append(FileLink(url=_pdf_view_url(pdf_urls[0]), format="Application/pdf"))
    if html_url:
        file_links.append(FileLink(url=html_url, format="text/html"))

    # Also capture mailto emails from page if author names match
    for a in soup.select('a[href^="mailto:"]'):
        email = a.get("href", "").replace("mailto:", "").split("?")[0].strip()
        label = a.get_text(" ", strip=True)
        if not email or "@" not in email:
            continue
        for author in authors:
            if not author.email and (
                email.lower().startswith(author.name.split()[0].lower())
                or (label and label.lower() in author.name.lower())
                or author == authors[0]
            ):
                # Prefer attaching unmatched mailto to first author without email
                if author.email:
                    continue
                author.email = email
                break
        else:
            if authors and not authors[0].email:
                authors[0].email = email

    page_doi = _first(_meta_values(soup, "citation_doi", "DC.Identifier.DOI", "dc.identifier"))
    if page_doi:
        page_doi = clean_doi(page_doi)

    return ArticleMeta(
        doi=page_doi or doi,
        title=title,
        abstract=_strip_jats(abstract),
        keywords=clean_keywords,
        authors=authors,
        journal=journal,
        pages=pages,
        volume=volume,
        issue=issue,
        year=year,
        month=month_name(month) if month else "",
        file_links=file_links,
        landing_url=landing_url,
        source="landing-page",
    )


async def _fetch_crossref(client: httpx.AsyncClient, doi: str) -> dict:
    url = f"https://api.crossref.org/works/{doi}"
    resp = await client.get(url, headers={"Accept": "application/json"})
    if resp.status_code != 200:
        return {}
    return resp.json().get("message") or {}


def _merge_crossref(meta: ArticleMeta, data: dict) -> ArticleMeta:
    if not data:
        return meta

    if not meta.title:
        titles = data.get("title") or []
        if titles:
            meta.title = titles[0]

    if not meta.abstract and data.get("abstract"):
        meta.abstract = _strip_jats(data["abstract"])

    if not meta.journal:
        container = data.get("container-title") or []
        if container:
            meta.journal = container[0].replace(" & ", " and ")

    if not meta.volume:
        meta.volume = str(data.get("volume") or "")
    if not meta.issue:
        meta.issue = str(data.get("issue") or "")
    if not meta.pages:
        meta.pages = str(data.get("page") or data.get("article-number") or "")

    if not meta.year or not meta.month:
        for key in ("published-print", "published-online", "published", "issued"):
            parts = ((data.get(key) or {}).get("date-parts") or [[]])[0]
            if parts:
                if not meta.year and len(parts) >= 1:
                    meta.year = str(parts[0])
                if not meta.month and len(parts) >= 2:
                    meta.month = month_name(parts[1])
                break

    if not meta.authors and data.get("author"):
        authors: list[Author] = []
        for item in data["author"]:
            given = (item.get("given") or "").strip()
            family = (item.get("family") or "").strip()
            name = f"{given} {family}".strip() or (item.get("name") or "").strip()
            workplace = None
            aff = item.get("affiliation") or []
            if aff:
                workplace = aff[0].get("name")
            email = item.get("email")
            authors.append(Author(name=name, email=email, workplace=workplace))
        meta.authors = authors
    else:
        # fill missing workplaces / emails from Crossref by position
        xref_authors = data.get("author") or []
        for idx, author in enumerate(meta.authors):
            if idx >= len(xref_authors):
                break
            item = xref_authors[idx]
            if not author.workplace:
                aff = item.get("affiliation") or []
                if aff:
                    author.workplace = aff[0].get("name")
            if not author.email and item.get("email"):
                author.email = item.get("email")

    if not meta.keywords:
        subjects = data.get("subject") or []
        meta.keywords = [s for s in subjects if s]

    # resource URL / links
    if not meta.file_links:
        resource = ((data.get("resource") or {}).get("primary") or {}).get("URL")
        links = []
        if resource:
            links.append(FileLink(url=resource, format="text/html"))
        for link in data.get("link") or []:
            url = link.get("URL")
            ctype = (link.get("content-type") or "").lower()
            if not url:
                continue
            if "pdf" in ctype:
                links.insert(0, FileLink(url=_pdf_view_url(url), format="Application/pdf"))
            else:
                links.append(FileLink(url=url, format=ctype or "text/html"))
        meta.file_links = links

    if not meta.landing_url:
        meta.landing_url = data.get("URL") or normalize_doi_url(meta.doi)

    if meta.source == "landing-page":
        meta.source = "landing-page+crossref"
    else:
        meta.source = "crossref"
    return meta


def _http_error_message(status_code: int, where: str) -> str:
    if status_code == 404:
        return f"{where} returned HTTP 404 (not found)"
    if status_code == 403:
        return f"{where} returned HTTP 403 (forbidden)"
    if status_code >= 500:
        return f"{where} returned HTTP {status_code} (server error)"
    if status_code >= 400:
        return f"{where} returned HTTP {status_code}"
    return f"{where} returned HTTP {status_code}"


async def extract_doi(
    doi: str,
    client: httpx.AsyncClient | None = None,
    timeout: float = 45.0,
) -> ArticleMeta:
    """Extract one DOI. Never raises — failures become ArticleMeta.error so batching can continue."""
    original = (doi or "").strip()
    doi = clean_doi(doi)
    if not doi:
        return ArticleMeta(doi="", input_ref=original, error="Empty DOI")

    owns_client = client is None
    if owns_client:
        client = httpx.AsyncClient(
            follow_redirects=True,
            timeout=timeout,
            headers={"User-Agent": USER_AGENT},
        )

    assert client is not None
    meta = ArticleMeta(doi=doi, input_ref=original or doi)

    try:
        # 1) Resolve DOI to landing page HTML
        try:
            resp = await client.get(
                normalize_doi_url(doi),
                headers={"Accept": "text/html,application/xhtml+xml"},
            )
            content_type = (resp.headers.get("content-type") or "").lower()
            meta.landing_url = str(resp.url)
            if resp.status_code >= 400:
                meta.error = _http_error_message(resp.status_code, "DOI landing page")
            elif "html" in content_type:
                meta = _extract_from_html(resp.text, str(resp.url), doi)
            else:
                meta.error = f"DOI landing page returned non-HTML content ({content_type or 'unknown'})"
        except httpx.TimeoutException:
            meta.error = "DOI landing page timed out"
        except httpx.HTTPError as exc:
            meta.error = f"DOI landing page not accessible: {exc}"
        except Exception as exc:  # noqa: BLE001
            meta.error = f"Landing page fetch failed: {exc}"

        # 2) Enrich / fill gaps from Crossref (also helps when landing page 404s)
        try:
            xref = await _fetch_crossref(client, doi)
            if xref:
                meta = _merge_crossref(meta, xref)
            elif not meta.title:
                # Keep prior landing-page error if present; otherwise note Crossref miss
                meta.error = meta.error or "DOI not found in Crossref"
        except httpx.TimeoutException:
            if not meta.title:
                meta.error = meta.error or "Crossref lookup timed out"
        except Exception as exc:  # noqa: BLE001
            if not meta.title:
                meta.error = meta.error or f"Crossref fetch failed: {exc}"

        if not meta.title:
            meta.error = meta.error or "Could not extract article metadata"
        else:
            meta.error = None

        # Ensure DOI URL casing prefers input DOI
        meta.doi = doi
        meta.input_ref = original or doi
        if meta.month and meta.month.isdigit():
            meta.month = month_name(meta.month)

        # Guarantee at least an HTML file link
        if not meta.file_links and meta.landing_url:
            meta.file_links = [FileLink(url=meta.landing_url, format="text/html")]

        return meta
    except Exception as exc:  # noqa: BLE001 — never break the batch
        return ArticleMeta(doi=doi, input_ref=original or doi, error=f"Unexpected error: {exc}")
    finally:
        if owns_client:
            await client.aclose()


async def extract_url(
    url: str,
    client: httpx.AsyncClient | None = None,
    timeout: float = 45.0,
) -> ArticleMeta:
    """Extract metadata directly from an article page URL (works even without a DOI)."""
    original = (url or "").strip()
    url = clean_url(url)
    if not url:
        return ArticleMeta(input_ref=original, error="Empty or invalid URL")

    owns_client = client is None
    if owns_client:
        client = httpx.AsyncClient(
            follow_redirects=True,
            timeout=timeout,
            headers={"User-Agent": USER_AGENT},
        )
    assert client is not None

    try:
        try:
            resp = await client.get(
                url,
                headers={"Accept": "text/html,application/xhtml+xml"},
            )
        except httpx.TimeoutException:
            return ArticleMeta(input_ref=original, landing_url=url, error="Article URL timed out")
        except httpx.HTTPError as exc:
            return ArticleMeta(
                input_ref=original,
                landing_url=url,
                error=f"Article URL not accessible: {exc}",
            )

        content_type = (resp.headers.get("content-type") or "").lower()
        landing = str(resp.url)
        if resp.status_code >= 400:
            return ArticleMeta(
                input_ref=original,
                landing_url=landing,
                error=_http_error_message(resp.status_code, "Article URL"),
            )
        if "html" not in content_type:
            return ArticleMeta(
                input_ref=original,
                landing_url=landing,
                error=f"Article URL returned non-HTML content ({content_type or 'unknown'})",
            )

        meta = _extract_from_html(resp.text, landing, doi="")
        meta.input_ref = original
        meta.landing_url = landing
        meta.source = "article-url"

        # Only trust explicit citation/DOI meta — never scrape random reference DOIs from the body
        page_doi = clean_doi(meta.doi) if meta.doi else ""
        if not page_doi:
            page_doi = _infer_doi_from_known_hosts(landing, resp.text)

        if page_doi:
            meta.doi = page_doi
            try:
                xref = await _fetch_crossref(client, page_doi)
                if xref:
                    meta = _merge_crossref(meta, xref)
                    meta.doi = page_doi
                    meta.input_ref = original
                    meta.source = "article-url+crossref"
            except Exception:
                pass
        else:
            meta.doi = ""

        if not meta.title:
            meta.error = meta.error or "Could not extract article metadata from URL"
        else:
            meta.error = None

        if meta.month and str(meta.month).isdigit():
            meta.month = month_name(meta.month)

        if not meta.file_links:
            meta.file_links = [FileLink(url=landing, format="text/html")]
        else:
            # Ensure the provided article URL is present as HTML link
            if not any(l.url.rstrip("/") == landing.rstrip("/") for l in meta.file_links):
                meta.file_links.append(FileLink(url=landing, format="text/html"))

        return meta
    except Exception as exc:  # noqa: BLE001
        return ArticleMeta(input_ref=original, landing_url=url, error=f"Unexpected error: {exc}")
    finally:
        if owns_client:
            await client.aclose()


def _infer_doi_from_known_hosts(landing_url: str, html: str) -> str:
    """Best-effort DOI inference for known hosts when citation_doi meta is missing."""
    landing = (landing_url or "").lower()
    # 50sea IJIST OJS articles often omit citation_doi; article id is in DC.Identifier / URL
    if "journal.50sea.com" in landing and "/ijist/" in landing:
        soup = BeautifulSoup(html or "", "lxml")
        article_id = _first(_meta_values(soup, "DC.Identifier", "dc.identifier"))
        if not article_id or not article_id.isdigit():
            m = re.search(r"/article/view/(\d+)", landing_url or "", flags=re.I)
            if m:
                article_id = m.group(1)
        if article_id and article_id.isdigit():
            return f"10.33411/IJIST/{article_id}"
    return ""


async def extract_item(
    value: str,
    client: httpx.AsyncClient | None = None,
    timeout: float = 45.0,
) -> ArticleMeta:
    """Extract from a DOI or an article URL."""
    value = (value or "").strip()
    if not value:
        return ArticleMeta(error="Empty input")
    url = clean_url(value)
    if url:
        if is_doi_org_url(url):
            return await extract_doi(url, client=client, timeout=timeout)
        return await extract_url(url, client=client, timeout=timeout)
    return await extract_doi(value, client=client, timeout=timeout)


async def extract_many(
    dois: list[str],
    concurrency: int = 5,
    progress_cb=None,
) -> list[ArticleMeta]:
    """Extract many DOIs and/or article URLs.

    progress_cb, if provided, is called with dict events:
      {"phase":"start","index":i,"doi":ref,"total":n}
      {"phase":"done","index":i,"doi":ref,"total":n,"meta":ArticleMeta}
    For backward compatibility it also accepts legacy callbacks of form (index, meta).
    """
    sem = asyncio.Semaphore(concurrency)
    results: list[ArticleMeta | None] = [None] * len(dois)
    total = len(dois)

    def _emit(event: dict) -> None:
        if not progress_cb:
            return
        try:
            progress_cb(event)
        except TypeError:
            # Legacy signature: progress_cb(index, meta)
            if event.get("phase") == "done":
                progress_cb(event["index"], event["meta"])

    async with httpx.AsyncClient(
        follow_redirects=True,
        timeout=45.0,
        headers={"User-Agent": USER_AGENT},
    ) as client:

        async def worker(idx: int, ref: str) -> None:
            async with sem:
                _emit({"phase": "start", "index": idx, "doi": ref, "total": total})
                try:
                    meta = await extract_item(ref, client=client)
                except Exception as exc:  # noqa: BLE001 — skip bad items, continue list
                    meta = ArticleMeta(input_ref=ref, error=f"Skipped due to error: {exc}")
                results[idx] = meta
                _emit(
                    {
                        "phase": "done",
                        "index": idx,
                        "doi": ref,
                        "total": total,
                        "meta": meta,
                    }
                )

        await asyncio.gather(*(worker(i, d) for i, d in enumerate(dois)), return_exceptions=False)

    return [
        r if r is not None else ArticleMeta(input_ref=dois[i], error="Unknown failure")
        for i, r in enumerate(results)
    ]
